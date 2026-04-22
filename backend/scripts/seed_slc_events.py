"""Seed SLC calendar submissions from the Auxiliary Services spreadsheet.

Reads FY26 events (April + May 2026) from AuxServices/Calendar Spreadsheet1.xlsx
and inserts them as Submissions with Target_Newsletter="none" and
Show_In_SLC_Calendar=True. Idempotent: skips any submission where the
Submitter_Email + Original_Headline + Requested_Date combo already exists.

Exists as a one-off exploration seed for the feature/slc-calendar-exploration
branch — not wired into production seed.py.
"""

import asyncio
import re
import sys
import warnings
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.db.engine import async_session_factory
from app.models.submission import Submission, SubmissionScheduleRequest

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SPREADSHEET = REPO_ROOT / "AuxServices" / "Calendar Spreadsheet1.xlsx"
SEED_EMAIL = "auxservices-seed@example.edu"
SEED_NAME = "Auxiliary Services (spreadsheet import)"


def _parse_cell_date(value) -> date | None:
    """Return a date for a cell that's either a datetime, date, or a m/d range string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        # Strip trailing "?" and whitespace, keep only the start of a range.
        cleaned = value.strip().rstrip("?").strip()
        cleaned = cleaned.split("-")[0].strip()
        m = re.match(r"^(\d{1,2})/(\d{1,2})$", cleaned)
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            # FY26 Apr–May data all falls in calendar-year 2026.
            return date(2026, month, day)
    return None


def _build_body(event: str, location: str | None, sponsor: str | None,
                start_time: str | None, ticketed: str | None) -> str:
    parts = [f"Event: {event}"]
    if location:
        parts.append(f"Location: {location}")
    if sponsor:
        parts.append(f"Sponsor: {sponsor}")
    if start_time:
        parts.append(f"Start Time: {start_time}")
    if ticketed:
        parts.append(f"Ticketed: {ticketed}")
    return "\n".join(parts)


async def seed_from_spreadsheet(spreadsheet_path: Path) -> int:
    if not spreadsheet_path.exists():
        print(f"Spreadsheet not found: {spreadsheet_path}", file=sys.stderr)
        return 0

    wb = load_workbook(spreadsheet_path, data_only=True)
    if "FY26" not in wb.sheetnames:
        print("FY26 sheet not found in spreadsheet", file=sys.stderr)
        return 0

    ws = wb["FY26"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    inserted = 0
    async with async_session_factory() as session:
        for row in rows:
            if not row or len(row) < 3:
                continue
            date_cell, start_time, event, location, sponsor, ticketed = (row + (None,) * 6)[:6]
            if not event or not str(event).strip():
                continue
            event_title = str(event).strip()
            event_date = _parse_cell_date(date_cell)
            if event_date is None:
                # Skip rows without a resolvable date.
                continue
            if event_date.year != 2026:
                # One known row has a 2024 typo for the Boise Commencement.
                continue
            if event_date.month not in (4, 5):
                continue

            # Idempotency check
            existing = await session.execute(
                select(Submission)
                .join(Submission.Schedule_Requests)
                .where(
                    Submission.Submitter_Email == SEED_EMAIL,
                    Submission.Original_Headline == event_title,
                    SubmissionScheduleRequest.Requested_Date == event_date,
                )
            )
            if existing.scalar_one_or_none():
                continue

            body = _build_body(
                event_title,
                str(location).strip() if location else None,
                str(sponsor).strip() if sponsor else None,
                str(start_time).strip() if start_time else None,
                str(ticketed).strip() if ticketed else None,
            )

            submission = Submission(
                Category="slc_event",
                Target_Newsletter="none",
                Original_Headline=event_title,
                Original_Body=body,
                Submitter_Name=SEED_NAME,
                Submitter_Email=SEED_EMAIL,
                Show_In_SLC_Calendar=True,
                Event_Classification=None,
                Status="approved",
            )
            session.add(submission)
            await session.flush()

            session.add(
                SubmissionScheduleRequest(
                    Submission_Id=submission.Id,
                    Requested_Date=event_date,
                    Repeat_Count=1,
                    Recurrence_Type="once",
                    Recurrence_Interval=1,
                    Excluded_Dates=[],
                )
            )
            inserted += 1

        await session.commit()

    return inserted


async def main() -> None:
    count = await seed_from_spreadsheet(SPREADSHEET)
    print(f"Inserted {count} SLC events from {SPREADSHEET.name}")


if __name__ == "__main__":
    asyncio.run(main())
