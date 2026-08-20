"""Import Senior Leadership Council calendar events from Excel workbooks.

The SLC calendar is backed by ``Submission`` records whose newsletter target is
``none``. This module translates the lightweight fiscal-year workbooks used by
Auxiliary Services into those records while preserving source context and date
uncertainty. Parsing stays separate from persistence so every import can be
previewed and tested before it changes a database.

Workbook date cells may be native Excel dates, single month/day strings, or
date ranges. A trailing question mark is treated as tentative rather than as
an instruction to invent certainty. Alternative dates joined with "or" are
reported for human review and are not imported automatically.
"""

from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.submission import Submission, SubmissionScheduleRequest

IMPORT_EMAIL = "slc-calendar-import@uidaho.edu"
IMPORT_NAME = "SLC calendar spreadsheet import"
IMPORT_KEY_PREFIX = "SLC spreadsheet import key:"
FISCAL_YEAR_SHEET_PATTERN = re.compile(r"^FY(?P<year>\d{2})$", re.IGNORECASE)
DATE_TOKEN_PATTERN = r"\d{1,2}/\d{1,2}(?:/\d{2,4})?"


class DateParseError(ValueError):
    """Raised when an event date cannot be mapped safely to calendar dates."""

    def __init__(self, message: str, code: str = "invalid_date") -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class ImportSchedule:
    """One calendar occurrence or contiguous date range from a workbook row."""

    start_date: date
    end_date: date | None = None

    @property
    def effective_end_date(self) -> date:
        return self.end_date or self.start_date


@dataclass(frozen=True)
class SLCEventImportRecord:
    """A validated event row ready to become an SLC-only submission."""

    sheet_name: str
    row_number: int
    title: str
    source_date: str
    schedules: tuple[ImportSchedule, ...]
    tentative: bool
    start_time: str | None = None
    location: str | None = None
    sponsor: str | None = None
    detail_label: str | None = None
    detail_value: str | None = None

    @property
    def import_key(self) -> str:
        payload = {
            "title": self.title.casefold(),
            "source_date": self.source_date.casefold(),
            "schedules": [
                [schedule.start_date.isoformat(), schedule.effective_end_date.isoformat()]
                for schedule in self.schedules
            ],
            "start_time": (self.start_time or "").casefold(),
            "location": (self.location or "").casefold(),
            "sponsor": (self.sponsor or "").casefold(),
        }
        encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode()
        return hashlib.sha256(encoded).hexdigest()[:24]


@dataclass(frozen=True)
class ImportIssue:
    """A source event that was intentionally omitted from an import plan."""

    sheet_name: str
    row_number: int
    title: str
    source_date: str
    code: str
    message: str


@dataclass
class SLCEventImportPlan:
    """Validated records and review items produced from one workbook."""

    workbook_name: str
    records: list[SLCEventImportRecord] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)

    def issue_counts(self) -> dict[str, int]:
        counts: dict[str, int] = {}
        for issue in self.issues:
            counts[issue.code] = counts.get(issue.code, 0) + 1
        return counts


@dataclass(frozen=True)
class ImportResult:
    """Persistence result for a validated import plan."""

    inserted: int
    existing: int
    would_insert: int


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, time):
        hour = value.hour % 12 or 12
        minute = f":{value.minute:02d}" if value.minute else ""
        suffix = "A" if value.hour < 12 else "P"
        return f"{hour}{minute}{suffix}"
    text = str(value).strip()
    return text or None


def _fiscal_year_bounds(fiscal_year: int) -> tuple[date, date]:
    return date(fiscal_year - 1, 7, 1), date(fiscal_year, 6, 30)


def _parse_date_token(token: str, fiscal_year: int) -> date:
    parts = [int(part) for part in token.split("/")]
    month, day = parts[:2]
    if len(parts) == 3:
        year = parts[2]
        if year < 100:
            year += 2000
    else:
        year = fiscal_year - 1 if month >= 7 else fiscal_year

    try:
        parsed = date(year, month, day)
    except ValueError as exc:
        raise DateParseError(f"Invalid calendar date: {token}") from exc

    fiscal_start, fiscal_end = _fiscal_year_bounds(fiscal_year)
    if not fiscal_start <= parsed <= fiscal_end:
        raise DateParseError(
            f"Date {parsed.isoformat()} falls outside FY{fiscal_year % 100:02d}",
            code="outside_fiscal_year",
        )
    return parsed


def parse_event_date(
    value: Any,
    fiscal_year: int,
) -> tuple[tuple[ImportSchedule, ...], bool, str]:
    """Parse one workbook date cell without guessing among alternatives."""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        fiscal_start, fiscal_end = _fiscal_year_bounds(fiscal_year)
        if not fiscal_start <= value <= fiscal_end:
            raise DateParseError(
                f"Date {value.isoformat()} falls outside FY{fiscal_year % 100:02d}",
                code="outside_fiscal_year",
            )
        return (ImportSchedule(value),), False, value.strftime("%-m/%-d/%Y")

    text = _clean_text(value)
    if not text:
        raise DateParseError("Event has no date", code="missing_date")

    tentative = "?" in text
    normalized = text.replace("?", "").strip()
    normalized = normalized.replace("–", "-").replace("—", "-")

    if re.search(r"\bor\b", normalized, flags=re.IGNORECASE):
        raise DateParseError(
            "Alternative dates joined with 'or' require confirmation",
            code="ambiguous_date",
        )

    and_parts = re.split(r"\s+and\s+", normalized, flags=re.IGNORECASE)
    if len(and_parts) > 1:
        if not all(re.fullmatch(DATE_TOKEN_PATTERN, part.strip()) for part in and_parts):
            raise DateParseError("Could not parse the dates joined with 'and'")
        schedules = tuple(
            ImportSchedule(_parse_date_token(part.strip(), fiscal_year))
            for part in and_parts
        )
        return schedules, tentative, text

    range_match = re.fullmatch(
        rf"(?P<start>{DATE_TOKEN_PATTERN})\s*-\s*(?P<end>{DATE_TOKEN_PATTERN})",
        normalized,
    )
    if range_match:
        start = _parse_date_token(range_match.group("start"), fiscal_year)
        end = _parse_date_token(range_match.group("end"), fiscal_year)
        if end < start:
            raise DateParseError("Date range ends before it starts")
        return (ImportSchedule(start, end),), tentative, text

    if re.fullmatch(DATE_TOKEN_PATTERN, normalized):
        parsed = _parse_date_token(normalized, fiscal_year)
        return (ImportSchedule(parsed),), tentative, text

    raise DateParseError(f"Unrecognized date format: {text}")


def _schedule_intersects_window(
    schedule: ImportSchedule,
    from_date: date,
    through_date: date | None,
) -> bool:
    if schedule.effective_end_date < from_date:
        return False
    if through_date and schedule.start_date > through_date:
        return False
    return True


def load_import_plan(
    workbook_path: Path,
    *,
    sheet_names: Iterable[str] | None = None,
    from_date: date,
    through_date: date | None = None,
    include_tentative: bool = True,
) -> SLCEventImportPlan:
    """Read fiscal-year worksheets and return a non-mutating import preview."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    requested_sheets = set(sheet_names or [])
    available_fiscal_sheets = [
        name for name in workbook.sheetnames if FISCAL_YEAR_SHEET_PATTERN.fullmatch(name)
    ]
    selected_sheets = (
        [name for name in available_fiscal_sheets if name in requested_sheets]
        if requested_sheets
        else available_fiscal_sheets
    )
    missing_sheets = requested_sheets.difference(workbook.sheetnames)
    if missing_sheets:
        missing = ", ".join(sorted(missing_sheets))
        raise ValueError(f"Workbook does not contain requested sheet(s): {missing}")
    if not selected_sheets:
        raise ValueError("Workbook has no fiscal-year sheets named like FY27")

    plan = SLCEventImportPlan(workbook_name=workbook_path.name)
    for sheet_name in selected_sheets:
        fiscal_year_match = FISCAL_YEAR_SHEET_PATTERN.fullmatch(sheet_name)
        assert fiscal_year_match is not None
        fiscal_year = 2000 + int(fiscal_year_match.group("year"))
        worksheet = workbook[sheet_name]
        headers = [_clean_text(cell.value) for cell in worksheet[1][:6]]
        if len(headers) < 5 or headers[:5] != [
            "Date",
            "Start Time",
            "Event",
            "Location",
            "Sponsor",
        ]:
            raise ValueError(
                f"{sheet_name} must start with Date, Start Time, Event, Location, Sponsor"
            )
        detail_label = headers[5] if len(headers) > 5 else None
        if detail_label and detail_label.casefold() == "catatory":
            detail_label = "Category"

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=6, values_only=True),
            start=2,
        ):
            values = tuple(row) + (None,) * (6 - len(row))
            title = _clean_text(values[2])
            if not title:
                continue
            source_date = _clean_text(values[0]) or ""
            try:
                schedules, tentative, source_date = parse_event_date(
                    values[0], fiscal_year
                )
            except DateParseError as exc:
                plan.issues.append(
                    ImportIssue(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        title=title,
                        source_date=source_date,
                        code=exc.code,
                        message=str(exc),
                    )
                )
                continue

            schedules_in_window = tuple(
                schedule
                for schedule in schedules
                if _schedule_intersects_window(schedule, from_date, through_date)
            )
            if not schedules_in_window:
                last_date = max(schedule.effective_end_date for schedule in schedules)
                code = "before_window" if last_date < from_date else "after_window"
                plan.issues.append(
                    ImportIssue(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        title=title,
                        source_date=source_date,
                        code=code,
                        message="Event falls outside the requested import window",
                    )
                )
                continue

            if tentative and not include_tentative:
                plan.issues.append(
                    ImportIssue(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        title=title,
                        source_date=source_date,
                        code="tentative_date",
                        message="Tentative dates were excluded by request",
                    )
                )
                continue

            plan.records.append(
                SLCEventImportRecord(
                    sheet_name=sheet_name,
                    row_number=row_number,
                    title=title,
                    source_date=source_date,
                    schedules=schedules_in_window,
                    tentative=tentative,
                    start_time=_clean_text(values[1]),
                    location=_clean_text(values[3]),
                    sponsor=_clean_text(values[4]),
                    detail_label=detail_label,
                    detail_value=_clean_text(values[5]),
                )
            )
    return plan


def _build_body(record: SLCEventImportRecord) -> str:
    parts = [f"Source date: {record.source_date}"]
    if record.tentative:
        parts.append("Date status: Tentative in source workbook")
    if record.start_time:
        parts.append(f"Start time: {record.start_time}")
    if record.location:
        parts.append(f"Location: {record.location}")
    if record.sponsor:
        parts.append(f"Sponsor: {record.sponsor}")
    if record.detail_label and record.detail_value:
        parts.append(f"{record.detail_label}: {record.detail_value}")
    parts.append(f"Source: {record.sheet_name}, row {record.row_number}")
    return "\n".join(parts)


def _build_notes(record: SLCEventImportRecord, workbook_name: str) -> str:
    return "\n".join(
        [
            f"{IMPORT_KEY_PREFIX} {record.import_key}",
            f"Source workbook: {workbook_name}",
            f"Source worksheet: {record.sheet_name}, row {record.row_number}",
        ]
    )


async def import_records(
    db: AsyncSession,
    plan: SLCEventImportPlan,
    *,
    apply_changes: bool,
) -> ImportResult:
    """Insert non-duplicate records, or report what an apply would insert."""
    inserted = 0
    existing = 0
    for record in plan.records:
        key_marker = f"{IMPORT_KEY_PREFIX} {record.import_key}"
        duplicate = await db.execute(
            select(Submission.Id)
            .where(
                Submission.Submitter_Email == IMPORT_EMAIL,
                Submission.Submitter_Notes.contains(key_marker),
            )
            .limit(1)
        )
        if duplicate.scalar_one_or_none() is not None:
            existing += 1
            continue
        if not apply_changes:
            continue

        submission = Submission(
            Category="slc_event",
            Target_Newsletter="none",
            Original_Headline=record.title,
            Original_Body=_build_body(record),
            Submitter_Name=IMPORT_NAME,
            Submitter_Email=IMPORT_EMAIL,
            Submitter_Notes=_build_notes(record, plan.workbook_name),
            Show_In_SLC_Calendar=True,
            Event_Classification=None,
            Status="pending_info" if record.tentative else "approved",
        )
        db.add(submission)
        await db.flush()

        for schedule in record.schedules:
            is_range = schedule.effective_end_date > schedule.start_date
            db.add(
                SubmissionScheduleRequest(
                    Submission_Id=submission.Id,
                    Requested_Date=schedule.start_date,
                    Repeat_Count=1,
                    Repeat_Note=(
                        "Tentative date from source workbook"
                        if record.tentative
                        else None
                    ),
                    Is_Flexible=record.tentative,
                    Recurrence_Type="date_range" if is_range else "once",
                    Recurrence_Interval=1,
                    Recurrence_End_Date=(
                        schedule.effective_end_date if is_range else None
                    ),
                    Excluded_Dates=[],
                )
            )
        inserted += 1

    if apply_changes:
        await db.commit()

    return ImportResult(
        inserted=inserted,
        existing=existing,
        would_insert=len(plan.records) - existing,
    )
